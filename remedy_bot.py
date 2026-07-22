# REMEDY_BOT IMPORTS ============== #

from telethon import events
import os
import asyncio
import re

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)
import pandas as pd
import matplotlib.pyplot as plt

from datetime import (
    datetime,
    timedelta,
    timezone
)

# ================= SOURCE GROUPS ================= #

# Comma-separated IDs supplied through environment configuration.
SOURCE_GROUPS = [
    int(value.strip())
    for value in os.getenv("REMEDY_SOURCE_GROUP_IDS", "").split(",")
    if value.strip()
]

SYNC_GROUPS = [
    int(value.strip())
    for value in os.getenv("REMEDY_SYNC_GROUP_IDS", "").split(",")
    if value.strip()
]

# ================= TARGET GROUPS ================= #

ROMH_FIBER_GROUP = int(os.getenv("REMEDY_OPERATIONS_CHAT_ID", "0"))
ROM_CIRCLE_LEADERS = int(os.getenv("REMEDY_LEADERS_CHAT_ID", "0"))

# ================= TIMEZONE ================= #

IST = timezone(
    timedelta(hours=5, minutes=30)
)

def now_ist():

    return datetime.now(IST)

# ================= AREA MAPPING ================= #

# Generic portfolio placeholders. Configure actual values through environment variables.
AREA_MAPPING = {
    "AREA_1": {"name": "AREA_1_LEADER", "id": int(os.getenv("AREA_1_LEADER_CHAT_ID", "0"))},
    "AREA_2": {"name": "AREA_2_LEADER", "id": int(os.getenv("AREA_2_LEADER_CHAT_ID", "0"))},
    "AREA_3": {"name": "AREA_3_LEADER", "id": int(os.getenv("AREA_3_LEADER_CHAT_ID", "0"))},
}

# ================= STORAGE ================= #

active_incidents = {}
# ================= STARTUP INCIDENT SYNC ================= #

async def sync_open_incidents(client):

    try:

        print("🔄 HARD SCANNING OPEN INCIDENTS")

        open_incidents = {}
        restored_incidents = set()

        async for dialog in client.iter_dialogs():

            group_name = "UNKNOWN"

            try:

                # ============================================
                # ONLY GROUPS / CHANNELS
                # ============================================

                if not (
                    dialog.is_group or
                    dialog.is_channel
                ):
                    continue

                group_name = str(dialog.name).upper()

                # ============================================
                # TARGET GROUP FILTER
                # ============================================

                ALLOWED_GROUPS = [
                    value.strip()
                    for value in os.getenv("REMEDY_ALLOWED_GROUP_NAMES", "").split(",")
                    if value.strip()
                ]

                group_name = (dialog.name or "").strip()

                if group_name not in ALLOWED_GROUPS:
                    continue

                print(
                    f"📌 SCANNING : {dialog.name}"
                )

                # ============================================
                # READ HISTORY
                # ============================================

                async for message in client.iter_messages(

                    dialog.entity,
                    limit=250,
                    reverse=False,

                ):

                    try:

                        # ====================================
                        # SKIP OLD MESSAGES
                        # ====================================

                        if not message.date:
                            continue

                        msg_time = (
                            message.date.astimezone(IST)
                        )

                        if (
                            now_ist() - msg_time
                        ).days > 2.5:

                            break

                        text = message.raw_text

                        if not text:
                            continue

                        upper = text.upper()

                        if "REMEDY" not in upper:
                            continue

                        incident = extract_incident(text)

                        if not incident:
                            continue

                        # ====================================
                        # RESTORED INCIDENT
                        # ====================================

                        if any(x in upper for x in [
                            "RESOLVED",
                            "RESTORED",
                            "UP TIME",
                            "UPTIME",
                            "RESOLUTION"
                        ]):

                            restored_incidents.add(
                                incident
                            )

                            continue

                        # ====================================
                        # OPEN INCIDENT
                        # ====================================

                        if (
                            "REMEDY NEW INCIDENT ALERT"
                            in upper
                        ):

                            if not valid_incident_type(upper):
                                continue

                            if not valid_section_ratio(upper):
                                continue

                            incident_type = (
                                extract_incident_type(text)
                                .replace(" (FC)", "")
                                .strip()
                            )

                            summary = extract_summary(text)

                            section_name = (
                                extract_section_name(text)
                            )

                            down_time = (
                                extract_down_time(text)
                            )

                            media_owner = extract_field(
                                text,
                                "Media Owner"
                            )

                            operator = extract_field(
                                text,
                                "Media Operator"
                            )

                            area = extract_area(upper)

                            tag = get_tag(area)

                            open_incidents[incident] = {

                                "incident_type": incident_type,
                                "summary": summary,
                                "section_name": section_name,
                                "down_time": down_time,
                                "media_owner": media_owner,
                                "operator": operator,
                                "area": area,
                                "tag": tag,
                                "last_followup": now_ist(),
                                "raw_text": text

                            }

                    except Exception:
                        continue

            except Exception as e:

                print(
                    f"❌ GROUP SCAN ERROR : "
                    f"{group_name} -> {e}"
                )

        # ============================================
        # FINAL ACTIVE INCIDENTS
        # ============================================

        active_incidents.clear()

        for inc, data in open_incidents.items():

            if inc in restored_incidents:
                continue
                
            latest_text = data.get("raw_text", "").upper()

            if "RESOLVED ALERT" in latest_text:
                continue

            active_incidents[inc] = data

        print(
            f"✅ HARD SCAN COMPLETE : "
            f"{len(active_incidents)} ACTIVE INCIDENTS"
        )

    except Exception as e:

        print(
            "❌ HARD SCAN ERROR :",
            e
        )
        
# ================= TIME FUNCTIONS ================= #

def format_duration(start_time):

    diff = now_ist() - start_time

    hrs = int(diff.total_seconds() // 3600)

    mins = int(
        (diff.total_seconds() % 3600) // 60
    )

    return f"{hrs}:{mins:02d} Hrs"


def format_duration_between(start, end):

    if not start or not end:
        return "NA"

    diff = end - start

    hrs = int(diff.total_seconds() // 3600)

    mins = int(
        (diff.total_seconds() % 3600) // 60
    )

    return f"{hrs}:{mins:02d} Hrs"

# ================= FIELD EXTRACTION ================= #

def extract_field(text, field):

    try:

        pattern = rf"{field}\s*:\s*(.*?)(?:,|$)"

        match = re.search(
            pattern,
            text,
            re.I
        )

        if not match:
            return "-"

        value = match.group(1).strip()

        if value == "":
            return "-"

        return value

    except:
        return "-"

def extract_incident(text):

    patterns = [

        r"Incident\s*[:-]\s*(INC\d+)",
        r"Incident ID\s*[:-]\s*(INC\d+)",
        r"CI ID\s*[:-]\s*(INC\d+)",
        r"TT\s*[:-]\s*(INC\d+)",
        r"\b(INC\d{6,})\b"
    ]

    for pattern in patterns:

        match = re.search(

            pattern,
            text,
            re.I
        )

        if match:

            return match.group(1).upper()

    return None

def extract_incident_type(text):

    upper = text.upper()

    if "LINEAR CONNECTIVITY FAILURE" in upper:
        return "Linear Connectivity Failure (FC)"

    elif "SECTION-CRI" in upper:
        return "Section-CRI (FC)"

    elif "MULTI SECTION" in upper:
        return "Multi Section (FC)"

    elif "RINGFAILURE" in upper:
        return "RingFailure (FC)"

    elif "RING FAILURE" in upper:
        return "RingFailure (FC)"

    elif "SECTIONISOLATION" in upper:
        return "SectionIsolation (FC)"

    elif "SECTION ISOLATION" in upper:
        return "SectionIsolation (FC)"

    elif "SECTION " in upper:
        return "Section (FC)"

    return "-"
    
def extract_summary(text):

    match = re.search(

        r"Summary\s*:\s*(.*?)(?:,\s*Down Since:|$)",

        text,
        re.I | re.S
    )

    return (

        match.group(1).strip()
        if match else "-"
    )

def extract_section_name(text):

    try:

        upper = text.upper()

        # ================= SUMMARY LINE ================= #

        summary_match = re.search(

            r"Summary:\s*(.*?)(?:, Down Since:|$)",

            text,
            re.I | re.S
        )

        if not summary_match:
            return "-"

        summary = summary_match.group(1)

        # ================= SECTION EXTRACTION ================= #

        links = re.findall(

            r"([A-Z0-9_-]+)/[A-Z0-9_-]+/[0-9]+:"
            r"([A-Z0-9_-]+)/[A-Z0-9_-]+/[0-9]+",

            summary,
            re.I
        )

        if not links:
            return "-"

        sections = []

        for a, b in links:

            sec = f"{a.upper()}-{b.upper()}"

            if sec not in sections:
                sections.append(sec)

        return " & ".join(sections)

    except:
        return "-"

# ================= TIME EXTRACTION ================= #

def extract_down_time(text):

    match = re.search(

        r"Down Since\s*[:-]\s*"
        r"(\d+/\d+/\d+\s+\d+:\d+:\d+\s+[AP]M)",

        text,
        re.I
    )

    if match:

        return datetime.strptime(

            match.group(1),
            "%m/%d/%Y %I:%M:%S %p"

        ).replace(tzinfo=IST)

    return now_ist()

def extract_up_time(text):

    patterns = [

        r"Up Time\s*[:-]\s*"
        r"(\d+/\d+/\d+\s+\d+:\d+:\d+\s+[AP]M)",

        r"Uptime\s*[:-]\s*"
        r"(\d+/\d+/\d+\s+\d+:\d+:\d+\s+[AP]M)",

        r"Resolved Date\s*[:-]\s*"
        r"(\d+/\d+/\d+\s+\d+:\d+:\d+\s+[AP]M)"
    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            text,
            re.I
        )

        if match:

            return datetime.strptime(

                match.group(1),
                "%m/%d/%Y %I:%M:%S %p"

            ).replace(tzinfo=IST)

    return None

# ================= AREA ================= #

def extract_area(text_upper):

    for area in AREA_MAPPING:

        if area in text_upper:
            return area

    return "NA"

def get_tag(area):

    if area in AREA_MAPPING:

        user = AREA_MAPPING[area]

        return (
            f'<a href="tg://user?id={user["id"]}">'
            f'{user["name"]}'
            f'</a>'
        )

    return "-"

# ================= TT EXTRACTION ================= #

# ================= TT EXTRACTION ================= #

def extract_tt_blocks(text):

    tt_blocks = []

    try:

        # ================= NSA TT BLOCK ================= #

        tt_section_match = re.search(

            r"NSA TT Details\s*:(.*)",

            text,
            re.I | re.S
        )

        if not tt_section_match:
            return []

        tt_section = tt_section_match.group(1)

        # ================= LINE EXTRACTION ================= #

        lines = [

            line.strip()

            for line in tt_section.splitlines()

            if "INC" in line.upper()
        ]

        count = 1

        for line in lines:

            try:

                # ================= INCIDENT ================= #

                inc_match = re.search(

                    r"(INC\d+)",

                    line,
                    re.I
                )

                if not inc_match:
                    continue

                inc = inc_match.group(1)

                # ================= SUMMARY ================= #

                summary_match = re.search(

                    r"INC\d+\s+([A-Z0-9/_-]+:[A-Z0-9/_-]+)",

                    line,
                    re.I
                )

                if not summary_match:
                    continue

                summary = summary_match.group(1)

                # ================= LABEL ================= #

                if count == 1:
                    label = "1st Incident"

                elif count == 2:
                    label = "2nd Incident"

                elif count == 3:
                    label = "3rd Incident"

                else:
                    label = f"{count}th Incident"

                # ================= SECTION ================= #

                sec_match = re.findall(

                    r"([A-Z0-9_-]+)/[A-Z0-9_-]+/[0-9]+:"
                    r"([A-Z0-9_-]+)/[A-Z0-9_-]+/[0-9]+",

                    summary,
                    re.I
                )

                section = "-"

                if sec_match:

                    a, b = sec_match[0]

                    section = (
                        f"{a.upper()}-{b.upper()}"
                    )

                # ================= TIME ================= #

                time_match = re.search(

                    r"([A-Z][a-z]{2}\s+[A-Z][a-z]{2}\s+\d+\s+\d+:\d+:\d+\s+\d{4})",

                    line
                )

                if time_match:

                    try:

                        dt = datetime.strptime(

                            time_match.group(1),
                            "%a %b %d %H:%M:%S %Y"

                        ).replace(tzinfo=IST)

                        downtime = dt.strftime(
                            "%d/%m/%Y %I:%M:%S %p"
                        )

                        duration = format_duration(dt)

                    except:

                        downtime = "-"
                        duration = "-"

                else:

                    downtime = "-"
                    duration = "-"

                tt_blocks.append({

                    "label": label,

                    "incident": inc,

                    "summary": summary,

                    "section": section,

                    "downtime": downtime,

                    "duration": duration
                })

                count += 1

            except Exception as e:

                print("TT LINE ERROR :", e)

    except Exception as e:

        print("TT ERROR :", e)

    return tt_blocks
# ================= VALIDATIONS ================= #

def valid_incident_type(text_upper):

    return any(x in text_upper for x in [

        "LINEAR CONNECTIVITY FAILURE",
        "SECTION ",
        "SECTION-CRI",
        "MULTI SECTION",
        "RINGFAILURE",
        "RING FAILURE",
        "SECTIONISOLATION",
        "SECTION ISOLATION"
    ])

def valid_section_ratio(text_upper):

    # ONLY apply ratio logic
    # for Section / Section-CRI

    if (
        "SECTION " not in text_upper and
        "SECTION-CRI" not in text_upper
    ):
        return True

    ratio_match = re.search(

        r"(\d+)/(\d+)\s+LINK DOWN",
        text_upper
    )

    # If no ratio found
    # allow incident

    if not ratio_match:
        return True

    down = int(ratio_match.group(1))
    total = int(ratio_match.group(2))

    if total == 0:
        return False

    percent = (down / total) * 100

    return percent >= 50
# ================= REGISTER HANDLER ================= #

def register_remedy_handler(client):

    @client.on(events.NewMessage())
    async def remedy_handler(event):

        try:

            text = event.raw_text

            upper = text.upper()

            # ================= REMOVE URL ================= #

            text = re.sub(
                r'https?://\S+',
                '',
                text
            )

            # ================= VALIDATIONS ================= #

            if "REMEDY" not in upper:
                return

            if not valid_incident_type(upper):
                return

            if not valid_section_ratio(upper):
                return

            # ================= EXTRACTIONS ================= #

            incident = extract_incident(text)

            if not incident:
                return

            incident_type = extract_incident_type(text)

            summary = extract_summary(text)

            section_name = extract_section_name(text)

            down_time = extract_down_time(text)

            up_time = extract_up_time(text)

            media_owner = extract_field(
                text,
                "Media Owner"
            )

            operator = extract_field(
                text,
                "Media Operator"
            )

            area = extract_area(upper)

            tag = get_tag(area)

            tt_blocks = extract_tt_blocks(text)

            # ================= OPEN INCIDENT ================= #

            if "REMEDY NEW INCIDENT ALERT" in upper:

                if incident in active_incidents:
                    return

                duration = format_duration(
                    down_time
                )

                msg = (

                    f"🚨 <b>Open Incident Alert</b> 🚨\n\n"

                    f"📅 <b>Date</b> - "
                    f"{now_ist().strftime('%d-%b-%y')}\n\n"

                    f"🎯 <b>Incident Type</b> - "
                    f"{incident_type}\n\n"

                    f"🆔 <b>Incident</b> - "
                    f"{incident}\n\n"

                    f"📄 <b>Summary</b> -\n"
                    f"{summary}\n\n"

                    f"📍 <b>Section Name</b> -\n"
                    f"{section_name}\n\n"

                    f"⏱ <b>Down Since</b> -\n"
                    f"{down_time.strftime('%d/%m/%Y %I:%M:%S %p')}\n\n"

                    f"🟢 <b>Up Time</b> -\n"
                    f"Awaited\n\n"

                    f"⌛ <b>Duration</b> -\n"
                    f"{duration}\n\n"

                    f"📡 <b>Media Owner</b> -\n"
                    f"{media_owner}\n\n"

                    f"🛠 <b>Operator</b> -\n"
                    f"{operator}\n\n"
                )

                # ================= TT BLOCK ================= #

                for tt in tt_blocks:

                    msg += (

                        f"🔹 <b>{tt['label']}</b> -\n"
                        f"{tt['incident']}\n\n"

                        f"📄 <b>{tt['label']} Summary</b> -\n"
                        f"{tt['summary']}\n\n"

                        f"📍 <b>{tt['label']} Section</b> -\n"
                        f"{tt['section']}\n\n"

                        f"⏱ <b>Downtime</b> -\n"
                        f"{tt.get('downtime', '-')}\n\n"

                        f"⌛ <b>Duration</b> -\n"
                        f"{tt.get('duration', '-')}\n\n"
                     )

                msg += (

                        f"📌 <b>Area</b> -\n"
                        f"{area}\n\n"

                        f"👤 <b>Tag</b> -\n"
                        f"{tag}"
                )

                #await client.send_message(

                    #TARGET_GROUP,
                    #msg,
                    #parse_mode="html"
                #)

                active_incidents[incident] = {

                    "incident_type": incident_type,
                    "summary": summary,
                    "section_name": section_name,
                    "down_time": down_time,
                    "media_owner": media_owner,
                    "operator": operator,
                    "area": area,
                    "tag": tag,
                    "tt_blocks": tt_blocks,
                    "last_followup": now_ist()
                }

                print(
                    f"✅ OPEN INCIDENT : {incident}"
                     )
            # ================= RESTORED ================= #

            elif any(x in upper for x in ["RESOLVED","RESTORED","UP TIME","UPTIME","RESOLUTION"]):

                if incident not in active_incidents:
                    return

                data = active_incidents[incident]

                duration = format_duration_between(

                    data["down_time"],
                    up_time
                )

                msg = (

                    f"✅ <b>Incident Restored</b> ✅\n\n"

                    f"🎯 <b>Incident Type</b> - "
                    f"{data['incident_type']}\n\n"

                    f"🆔 <b>Incident</b> - "
                    f"{incident}\n\n"

                    f"📄 <b>Summary</b> -\n"
                    f"{data['summary']}\n\n"

                    f"📍 <b>Section Name</b> -\n"
                    f"{data['section_name']}\n\n"

                    f"⏱ <b>Down Since</b> -\n"

                    f"{data['down_time'].strftime('%d/%m/%Y %I:%M:%S %p')}\n\n"

                    f"🟢 <b>Up Time</b> -\n"

                    f"{up_time.strftime('%d/%m/%Y %I:%M:%S %p') if up_time else 'NA'}\n\n"

                    f"⌛ <b>Total Duration</b> -\n"
                    f"{duration}\n\n"

                    f"📡 <b>Media Owner</b> -\n"
                    f"{data['media_owner']}\n\n"

                    f"🛠 <b>Operator</b> -\n"
                    f"{data['operator']}\n\n"

                    f"📌 <b>Area</b> -\n"
                    f"{data['area']}\n\n"

                    f"👤 <b>Tag</b> -\n"
                    f"{data['tag']}"
                )

                #await client.send_message(

                    #TARGET_GROUP,
                    #msg,
                    #parse_mode="html"
                #)

                del active_incidents[incident]

                print(
                    f"✅ RESTORED : {incident}"
                )

        except Exception as e:

            print("❌ REMEDY ERROR :", e)

# ================= FOLLOWUP CHECKER ================= #

#async def remedy_followup_checker(client):

#    while True:

#        try:

#            now = now_ist()

#            for inc, data in list(
#                active_incidents.items()
#            ):

#                diff = now - data["down_time"]

                # ONLY >4 HRS

#                if diff.total_seconds() < 4 * 3600:
#                    continue

#                # EVERY 2 HRS

#                if now - data["last_followup"] >= timedelta(hours=2):

#                    duration = format_duration(
#                        data["down_time"]
#                    )

#                    msg = (

#                        f"⏰ <b>Incident Follow-up (>4 Hrs)</b>\n\n"

#                        f"📅 <b>Date</b> - "
#                        f"{now_ist().strftime('%d-%b-%y')}\n\n"

#                        f"🎯 <b>Incident Type</b> - "
#                        f"{data['incident_type']}\n\n"

#                        f"🆔 <b>Incident</b> - "
#                        f"{inc}\n\n"

#                        f"📄 <b>Summary</b> -\n"
#                        f"{data['summary']}\n\n"

#                        f"📍 <b>Section Name</b> -\n"
#                        f"{data['section_name']}\n\n"

#                        f"⏱ <b>Down Since</b> -\n"
#                        f"{data['down_time'].strftime('%d/%m/%Y %I:%M:%S %p')}\n\n"

#                        f"⌛ <b>Running Duration</b> -\n"
#                        f"{duration}\n\n"

#                        f"📡 <b>Media Owner</b> -\n"
#                        f"{data['media_owner']}\n\n"

#                        f"🛠 <b>Operator</b> -\n"
#                        f"{data['operator']}\n\n"

#                        f"📌 <b>Area</b> -\n"
#                        f"{data['area']}\n\n"

#                        f"👤 <b>Tag</b> -\n"
#                        f"{data['tag']}"
#                    )

                    # ================= TT BLOCKS ================= #
#                    for tt in data.get("tt_blocks", []):
 #                       msg += (
#
 #                           f"🔹 <b>{tt['label']}</b> -\n"
  #                          f"{tt['incident']}\n\n"

   #                         f"📄 <b>{tt['label']} Summary</b> -\n"
    #                        f"{tt['summary']}\n\n"

     #                       f"📍 <b>{tt['label']} Section</b> -\n"
      #                      f"{tt['section']}\n\n"

       #                     f"⏱ <b>Downtime</b> -\n"
        #                    f"{tt.get('downtime', '-')}\n\n"

         #                   f"⌛ <b>Duration</b> -\n"
          #                  f"{tt.get('duration', '-')}\n\n"

          #              )

           #         msg += (
            #            f"📌 <b>Area</b> -\n"
             #           f"{data['area']}\n\n"

              #          f"👤 <b>Tag</b> -\n"
               #         f"{data['tag']}"

                #    )

                 #   await client.send_message(

                  #      TARGET_GROUP,
                   #     msg,
                    #    parse_mode="html"
                    #)

                    #data["last_followup"] = now

                    #print(
                     #   f"⏰ FOLLOWUP SENT : {inc}"
                    #)

#        except Exception as e:

#            print(
#                "❌ FOLLOWUP ERROR :",
#                e
#            )

#        await asyncio.sleep(600)

# ================= SUMMARY ================= #

async def summary_checker(client):

    while True:

        try:

            now = now_ist()

            # ================= NEXT 30-MIN SLOT ================= #

            next_minute = (
                ((now.minute // 15) + 1) * 15
            ) % 60

            next_hour = now.hour

            if next_minute == 0:
                next_hour = (next_hour + 1) % 24

            next_run = now.replace(

                hour=next_hour,
                minute=next_minute,
                second=0,
                microsecond=0
            )

            # midnight handling

            if next_run <= now:
                next_run += timedelta(days=1)

            sleep_seconds = (
                next_run - now
            ).total_seconds()

            print(
                f"📋 NEXT SUMMARY : "
                f"{next_run.strftime('%H:%M:%S')}"
            )

            await asyncio.sleep(
                sleep_seconds
            )

            # ================= SEND SUMMARY ================= #

            if active_incidents:

                await generate_remedy_summary(
                    client,
                    ROMH_FIBER_GROUP
                )

                print("📋 SUMMARY SENT")

        except Exception as e:

            print(
                "❌ SUMMARY ERROR :",
                e
            )

# ================= 2 HR LEADERS FLOW ================= #

async def leaders_summary_checker(client):

    while True:

        try:

            now = now_ist()

            next_hour = (
                ((now.hour // 2) + 1) * 2
            ) % 24

            next_run = now.replace(

                hour=next_hour,
                minute=0,
                second=0,
                microsecond=0
            )

            if next_run <= now:
                next_run += timedelta(days=1)

            sleep_seconds = (
                next_run - now
            ).total_seconds()

            print(
                f"📢 NEXT LEADERS FLOW : "
                f"{next_run.strftime('%H:%M:%S')}"
            )

            await asyncio.sleep(
                sleep_seconds
            )

            if active_incidents:

                await generate_remedy_summary(

                    client,
                    ROM_CIRCLE_LEADERS
                )

                print(
                    "✅ LEADERS FLOW SENT"
                )

        except Exception as e:

            print(
                "❌ LEADERS FLOW ERROR :",
                e
            )
# ================= EXCEL SUMMARY ================= #

async def generate_remedy_summary(
    client,
    target_group
):

    try:

        wb = Workbook()

        # =====================================================
        # SHEET 1 : OPEN INCIDENT SUMMARY
        # =====================================================

        ws1 = wb.active

        ws1.title = "Open Incident Summary"

        incident_types = [

            "RingFailure (FC)",
            "Linear Connectivity Failure (FC)",
            "SectionIsolation (FC)",
            "Section-CRI (FC)",
            "Multi Section (FC)",
            "Section (FC)"
        ]

        areas = list(AREA_MAPPING.keys())

        headers = ["Area"]

        for itype in incident_types:

            headers.append(f"{itype} <4 Hrs")
            headers.append(f"{itype} >4 Hrs")

        headers += [

            "Grand Total <4",
            "Grand Total >4",
            "Total"
        ]

        ws1.append(headers)

        # ================= SUMMARY DATA ================= #

        for area in areas:

            row = [area]

            total_lt4 = 0
            total_gt4 = 0

            for itype in incident_types:

                lt4 = 0
                gt4 = 0

                for inc, data in active_incidents.items():

                    if (
                        data["area"].upper() == area.upper()
                        and data["incident_type"] == itype
                    ):

                        diff = (
                            now_ist() -
                            data["down_time"]
                        ).total_seconds() / 3600

                        if diff < 4:
                            lt4 += 1
                        else:
                            gt4 += 1

                row.append(lt4)
                row.append(gt4)

                total_lt4 += lt4
                total_gt4 += gt4

            row += [

                total_lt4,
                total_gt4,
                total_lt4 + total_gt4
            ]

            ws1.append(row)

        # =====================================================
        # SHEET 2 : OPEN INCIDENT DETAILS
        # =====================================================

        ws2 = wb.create_sheet(
            title="Open Incident Details"
        )

        headers2 = [

            "Date",
            "Incident Type",
            "Incident ID",
            "Summary",
            "Section Name",
            "Down Since",
            "Uptime",
            "Duration",
            "Area",
            "Media Owner",
            "Operator"
        ]

        ws2.append(headers2)

        for inc, data in active_incidents.items():

            duration = format_duration(
                data["down_time"]
            )

            ws2.append([

                now_ist().strftime("%d-%b-%y"),

                data["incident_type"],

                inc,

                data["summary"],

                data["section_name"],

                data["down_time"].strftime(
                    "%d/%m/%Y %I:%M:%S %p"
                ),

                "-",

                duration,

                data["area"],

                data["media_owner"],

                data["operator"]
            ])

        # =====================================================
        # SHEET 3 : DETAILED TT BACKUP
        # =====================================================

        ws3 = wb.create_sheet(
            title="Detailed TT Backup"
        )

        headers3 = [

            "Incident ID",
            "Incident Type",

            "1st Incident",
            "1st Incident Summary",
            "1st Incident Section",
            "1st Downtime",
            "1st Duration",

            "2nd Incident",
            "2nd Incident Summary",
            "2nd Incident Section",
            "2nd Downtime",
            "2nd Duration"
        ]

        ws3.append(headers3)

        for inc, data in active_incidents.items():

            tt1 = {}
            tt2 = {}

            if len(data.get("tt_blocks", [])) >= 1:
                tt1 = data["tt_blocks"][0]

            if len(data.get("tt_blocks", [])) >= 2:
                tt2 = data["tt_blocks"][1]

            ws3.append([

                inc,

                data["incident_type"],

                tt1.get("incident", ""),
                tt1.get("summary", ""),
                tt1.get("section", ""),
                tt1.get("downtime", ""),
                tt1.get("duration", ""),

                tt2.get("incident", ""),
                tt2.get("summary", ""),
                tt2.get("section", ""),
                tt2.get("downtime", ""),
                tt2.get("duration", "")
            ])

        # =====================================================
        # FORMATTING
        # =====================================================

        fill = PatternFill(

            start_color="D9EAD3",
            end_color="D9EAD3",
            fill_type="solid"
        )

        for ws in [ws1, ws2, ws3]:

            for cell in ws[1]:

                cell.font = Font(
                    bold=True
                )

                cell.fill = fill

                cell.alignment = Alignment(

                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # =====================================================
        # AUTO WIDTH
        # =====================================================

        for ws in [ws1, ws2, ws3]:

            for col in ws.columns:

                max_length = 0

                column = col[0].column_letter

                for cell in col:

                    try:

                        if len(str(cell.value)) > max_length:

                            max_length = len(
                                str(cell.value)
                            )

                    except:
                        pass

                adjusted_width = min(
                    max_length + 5,
                    60
                )

                ws.column_dimensions[
                    column
                ].width = adjusted_width

        # =====================================================
        # SAVE
        # =====================================================

        file_name = (

            f"Open_Incident_Summary_"
            f"{now_ist().strftime('%d%m%Y_%H%M')}.xlsx"
        )

        wb.save(file_name)

        # =====================================================
        # SEND FILE
        # =====================================================
        # =====================================================
        # SEND SUMMARY MESSAGE
        # =====================================================

        summary_msg = (

            f"📊 <b>Open Incident Summary</b>\n\n"

            f"🕒 "
            f"{now_ist().strftime('%d-%b-%Y %I:%M %p')}\n\n"

            f"🔴 Total Open Incidents : "
            f"{len(active_incidents)}\n\n"

            f"📁 Excel backup attached"
        )

        await client.send_message(

            target_group,
            summary_msg,
            parse_mode="html"
        )

        # =====================================================
        # SEND EXCEL FILE
        # =====================================================

        await client.send_file(

            target_group,

            file_name,

            caption=(
                f"📁 Open Incident Backup\n"
                f"🕒 "
                f"{now_ist().strftime('%d-%b-%Y %I:%M %p')}"
            )
        )

        await send_summary_images(client,target_group)

        print(
            "✅ EXCEL SUMMARY SENT"
        )

    except Exception as e:

        print(
            "❌ EXCEL SUMMARY ERROR :",
            e
        )
async def send_summary_images(client,target_group):

    try:

        import pandas as pd
        import matplotlib.pyplot as plt
        import textwrap
        import math

        # =========================================================
        # SUMMARY IMAGE
        # =========================================================

        summary_data = []

        incident_types = [

            "RingFailure",
            "Linear Connectivity Failure",
            "SectionIsolation",
            "Section-CRI",
            "Multi Section",
            "Section"
        ]

        for area in AREA_MAPPING.keys():

            row = [area]

            grand_lt4 = 0
            grand_gt4 = 0

            for itype in incident_types:

                lt4 = 0
                gt4 = 0

                for inc, data in active_incidents.items():

                    dtype = (data["incident_type"].replace("(FC)", "").replace("  ", " ").strip().upper())

                    itype_clean = (itype.strip().upper())
                    
                    if (
                        data["area"] == area and
                        dtype == itype_clean
                    ):

                        hrs = (
                            now_ist() -
                            data["down_time"]
                        ).total_seconds() / 3600

                        if hrs < 4:
                            lt4 += 1
                        else:
                            gt4 += 1

                row.extend([lt4, gt4])

                grand_lt4 += lt4
                grand_gt4 += gt4

            row.extend([

                grand_lt4,
                grand_gt4,
                grand_lt4 + grand_gt4
            ])

            summary_data.append(row)

        columns = [

            "Area",

            "RF <4",
            "RF >4",

            "LCF <4",
            "LCF >4",

            "SI <4",
            "SI >4",

            "CRI <4",
            "CRI >4",

            "MS <4",
            "MS >4",

            "SEC <4",
            "SEC >4",

            "GT <4",
            "GT >4",

            "Total"
        ]

        df_summary = pd.DataFrame(
            summary_data,
            columns=columns
        )

        total_row = ["Total"]

        for col in columns[1:]:

            total_row.append(
                df_summary[col].sum()
            )

        df_summary.loc[len(df_summary)] = total_row

        # =========================================================
        # PROFESSIONAL SUMMARY IMAGE
        # =========================================================

        fig, ax = plt.subplots(
            figsize=(24, 8)
        )

        ax.axis("off")

        table = ax.table(

            cellText=df_summary.values,
            colLabels=df_summary.columns,

            cellLoc='center',
            loc='center'
        )

        table.auto_set_font_size(False)
        table.set_fontsize(11)

        for (row, col), cell in table.get_celld().items():

            cell.set_edgecolor("black")
            cell.set_linewidth(0.8)

            if row == 0:

                cell.set_facecolor("#d9c2e9")

                cell.set_text_props(

                    weight='bold',
                    fontsize=11,
                    color='black'
                )

                cell.set_height(0.075)

            elif row == len(df_summary):

                cell.set_facecolor("#ead1dc")

                cell.set_text_props(

                    weight='bold',
                    fontsize=11,
                    color='black'
                )

                cell.set_height(0.065)

            else:

                cell.set_text_props(

                    fontsize=10,
                    color='black'
                )

                cell.set_height(0.060)

        ax.set_title(

            f"Open Incident Summary ({now_ist().strftime('%d-%b-%y %I:%M %p')})",

            fontsize=18,
            weight='bold',
            pad=20
        )

        plt.savefig(

            "summary.png",

            bbox_inches='tight',
            dpi=300,
            pad_inches=0.3
        )

        plt.close()

        await client.send_file(

            target_group,
            "summary.png",

            caption="📊 Open Incident Summary"
        )

        # =========================================================
        # AREA WISE DETAILS
        # =========================================================

        area_data = {}

        for inc, data in active_incidents.items():

            area = data["area"]

            if area not in area_data:
                area_data[area] = []

            summary = textwrap.fill(

                str(data["summary"]),
                width=45
            )

            section = textwrap.fill(

                str(data["section_name"]),
                width=30
            )

            area_data[area].append([

                now_ist().strftime("%d-%b-%y"),

                data["incident_type"],

                inc,

                summary,

                section,

                data["down_time"].strftime(
                    "%d/%m/%Y %I:%M %p"
                ),

                "-",

                format_duration(
                    data["down_time"]
                ),

                area,

                data["media_owner"],

                data["operator"]
            ])

        # =========================================================
        # CREATE AREA IMAGES
        # =========================================================

        for area, incidents in area_data.items():

            chunks = [

                incidents[i:i+10]

                for i in range(
                    0,
                    len(incidents),
                    10
                )
            ]

            for idx, chunk in enumerate(chunks, start=1):

                df = pd.DataFrame(

                    chunk,

                    columns=[

                        "Date",
                        "Incident Type",
                        "Incident ID",
                        "Summary",
                        "Section Name",
                        "Down Since",
                        "Uptime",
                        "Duration",
                        "Area",
                        "Media Owner",
                        "Operator"
                    ]
                )

                fig_height = 7 + (len(df) * 0.55)

                fig, ax = plt.subplots(

                    figsize=(30, fig_height)
                )

                ax.axis("off")

                col_widths = [

                    0.07,
                    0.11,
                    0.12,
                    0.35,
                    0.20,
                    0.16,
                    0.06,
                    0.07,
                    0.08,
                    0.08,
                    0.08
                ]

                table = ax.table(

                    cellText=df.values,

                    colLabels=df.columns,

                    cellLoc='center',

                    colWidths=col_widths,

                    loc='center'
                )

                table.auto_set_font_size(False)

                table.set_fontsize(12)

                # =========================================================
                # CELL DESIGN
                # =========================================================

                for (row, col), cell in table.get_celld().items():

                    cell.set_edgecolor("black")
                    cell.set_linewidth(0.9)

                    if row == 0:

                        cell.set_facecolor("#d9c2e9")

                        cell.set_text_props(

                            weight='bold',
                            fontsize=14,
                            color='black'
                        )

                        cell.set_height(0.080)

                    else:

                        cell.set_text_props(

                            fontsize=13,
                            color='black'
                        )

                        cell.set_height(0.090)

                # =========================================================
                # TITLE
                # =========================================================

                if len(chunks) == 1:

                    title = (
                        f"{area} Open Incident Details"
                    )

                else:

                    title = (
                        f"{area} Open Incident Details "
                        f"(Part-{idx})"
                    )

                ax.set_title(

                    title,

                    fontsize=18,
                    weight='bold',
                    pad=20
                )

                image_name = (

                    f"{area}_{idx}.png"
                )

                plt.savefig(

                    image_name,

                    bbox_inches='tight',

                    dpi=300,

                    pad_inches=0.3
                )

                plt.close()

                tag = get_tag(area)

                await client.send_file(

                    target_group,

                    image_name,

                    caption=(

                        f"📍 {area} Open Incidents\n"
                        f"👤 {tag}"
                    ),

                    parse_mode="html"
                )

        print("✅ SUMMARY IMAGES SENT")

    except Exception as e:

        print(
            "❌ IMAGE ERROR :",
            e
        )
